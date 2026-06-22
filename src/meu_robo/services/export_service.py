from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from meu_robo.config import get_exports_dir
from meu_robo.repositories.vagas_repository import listar_vagas


def exportar_vagas_excel(filtros: dict | None = None) -> str:
    vagas = listar_vagas(filtros)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vagas"

    headers = [
        "ID",
        "URL",
        "Titulo extraido",
        "Empresa",
        "Localidade extraida",
        "Descricao",
        "Conteudo extraido em",
        "URL acessivel",
        "Erro extracao",
        "Plataforma",
        "Status",
        "Nota",
        "Recomendacao IA",
        "Justificativa IA",
        "Pontos positivos IA",
        "Pontos alerta IA",
        "Tecnologias IA",
        "Senioridade IA",
        "Analisada em",
        "Erro analise IA",
        "Titulo pesquisado",
        "Localidade pesquisada",
        "Encontrada em",
    ]
    sheet.append(headers)
    sheet.freeze_panes = "A2"

    for vaga in vagas:
        sheet.append(
            [
                vaga["id"],
                vaga["url"],
                vaga["titulo_vaga"],
                vaga["empresa"],
                vaga["localidade_extraida"],
                vaga["descricao_vaga"],
                vaga["conteudo_extraido_em"].strftime("%Y-%m-%d %H:%M:%S")
                if vaga["conteudo_extraido_em"]
                else "",
                vaga["url_acessivel"],
                vaga["erro_extracao"],
                vaga["plataforma"],
                vaga["status"],
                vaga["nota_aderencia"],
                vaga["recomendacao_ia"],
                vaga["justificativa_ia"],
                vaga["pontos_positivos_ia"],
                vaga["pontos_alerta_ia"],
                vaga["tecnologias_ia"],
                vaga["senioridade_ia"],
                vaga["analisada_em"].strftime("%Y-%m-%d %H:%M:%S")
                if vaga["analisada_em"]
                else "",
                vaga["erro_analise_ia"],
                vaga["titulo_busca"],
                vaga["localidade_busca"],
                vaga["encontrada_em"].strftime("%Y-%m-%d %H:%M:%S")
                if vaga["encontrada_em"]
                else "",
            ]
        )
        url_cell = sheet.cell(row=sheet.max_row, column=2)
        if vaga["url"]:
            url_cell.hyperlink = vaga["url"]
            url_cell.style = "Hyperlink"
            url_cell.font = Font(color="0563C1", underline="single")

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 80)

    filename = f"vagas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = get_exports_dir() / filename
    workbook.save(output_path)
    return str(output_path)
